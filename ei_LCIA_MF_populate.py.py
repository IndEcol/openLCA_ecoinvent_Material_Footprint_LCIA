# -*- coding: utf-8 -*-
"""

Last modified June 2021
@author: pauliuk
"""

# Script  ei_LCIA_MF_populate.py

# Import required libraries:
#%%

import openpyxl
import numpy as np
import os
import uuid
import json
import mf_Paths

#############################
#     Functions & Constants #
#############################
def CF_generate(mli,Val,dnames,duuid,dunit,ei_version_string):
    # create dictionary with characterisation factor

    if ei_version_string == '_ei_3_7_1':
        U_Mass = {
            "@type": "Unit",
            "@id": "20aadc24-a391-41cf-b340-3e4529f44bde",
            "name": "kg"}
        
        U_Energy = {
            "@type": "Unit",
            "@id": "52765a6c-3896-43c2-b2f4-c679acf13efe",
            "name": "MJ"}
        
        U_Volume = {
            "@type": "Unit",
            "@id": "1c3a9695-398d-4b1f-b07e-a8715b610f70",
            "name": "m3"}
        
        FP_Mass = {
            "@type": "FlowProperty",
            "@id": "93a60a56-a3c8-11da-a746-0800200b9a66",
            "name": "Mass",
            "categoryPath": [
                "Technical flow properties"]}
        
        FP_Energy = {
            "@type": "FlowProperty",
            "@id": "f6811440-ee37-11de-8a39-0800200c9a66",
            "name": "Energy",
            "categoryPath": [
                "Technical flow properties"]}
        
        FP_Volumne = {
            "@type": "FlowProperty",
            "@id": "93a60a56-a3c8-22da-a746-0800200c9a66",
            "name": "Volume",
            "categoryPath": [
                "Technical flow properties"]}

    CF = {}
    CF["@type"] = "ImpactFactor"
    CF["value"] = Val[mli]
    CF["flow"]  = {"@type": "Flow", "@id": duuid[mli],
    "name": dnames[mli],
    "categoryPath": [
        "Elementary flows",
        "Resource",
        "in ground"],
    "flowType": "ELEMENTARY_FLOW",
    "refUnit": dunit[mli]}
    if dunit[mli] == 'kg':
        CF["unit"]  = U_Mass
        CF["flowProperty"]  = FP_Mass
    elif dunit[mli] == 'MJ':
        CF["unit"]  = U_Energy
        CF["flowProperty"]  = FP_Energy                    
    elif dunit[mli] == 'm3':
        CF["unit"]  = U_Volume
        CF["flowProperty"]  = FP_Volumne
    else:
        None    
        
    return CF        

#################
#     MAIN      #
#################
# Set configuration data

ei_version_string = '_ei_3_7_1'
#ei_version_string = '_ei_3_8'

#%%
if ei_version_string == '_ei_3_7_1':
    tp   = mf_Paths.data_path_ei371
    MSn  = 'LCIA_Define_ecoinvent_3_7'
    MDn  = 'ecoinvent_3_7_Match'
    DN   = 411
if ei_version_string == '_ei_3_8':
    tp   = mf_Paths.data_path_ei38
    MSn  = 'LCIA_Define_ecoinvent_3_8'
    MDn  = 'ecoinvent_3_8_Match'
    DN   = 411
    
ScriptConfig = {}
ScriptConfig['Current_UUID'] = str(uuid.uuid4())

###################################################################################
#   Import data from masterfile                                                   #
###################################################################################   

# open master file
MasterFile  = openpyxl.load_workbook(os.path.join(mf_Paths.data_path_main,'Material_Footprint_LCIA_Master_V1.xlsx'),data_only=True)
# read LCIA indicator method uuids
MS = MasterFile[MSn]
mf_uuid = []
wf_uuid = []
for m in range(10,22):
    mf_uuid.append(MS.cell(m, 5).value)
for m in range(10,14):
    wf_uuid.append(MS.cell(m, 15).value)
# read master data
MD = MasterFile[MDn]
dnames  = []
duuid   = []
dselect = []   
dunit   = []
drmi    = []
dtmr    = []
for m in range(2,2+DN):
    dnames.append( MD.cell(m, 2).value)    
    duuid.append(  MD.cell(m, 4).value)    
    dselect.append(MD.cell(m,11).value)    
    dunit.append(  MD.cell(m,15).value)    
    drmi.append(   MD.cell(m,16).value)    
    dtmr.append(   MD.cell(m,21).value)     
# Tables with 1/0 flags to select individual factor for a given indicator    
MFSel = np.zeros((DN,6))
TFSel = np.zeros((DN,6))
WFSel = np.zeros((DN,4))    
for m in range(2,2+DN):
    for n in range(24,30):
        MFSel[m-2,n-24] = MD.cell(m,n).value
    for n in range(30,36):
        TFSel[m-2,n-30] = MD.cell(m,n).value
    for n in range(36,40):
        WFSel[m-2,n-36] = MD.cell(m,n).value
        
###################################################################################
#   Sort data into json files                                                     #
################################################################################### 
#%%
# loop over RMI files
for m in range(0,6):
    f_in = os.path.join(tp,'lcia_categories',mf_uuid[m]+'.json')
    with open(f_in, 'r+') as f:
        thisd = json.load(f)
        del thisd['impactFactors'][0] # delete the two factors that are still there from copying the files
        del thisd['impactFactors'][0]
        # add new impact factors from master data
        for mli in range(0,DN):
            if dselect[mli] != 1 and MFSel[mli,m] == 1: # add this value as impact/characterisation factor
                CF = CF_generate(mli,drmi,dnames,duuid,dunit,ei_version_string)
                # add new CF to json file:
                thisd['impactFactors'].append(CF)
        # wrap up and save    
        f.seek(0)        # reset file position to the beginning.
        json.dump(thisd, f, indent=4)
        f.truncate()     # remove remaining part
        f.close()

# loop over TMR files
for m in range(0,6):
    f_in = os.path.join(tp,'lcia_categories',mf_uuid[m+6]+'.json')
    with open(f_in, 'r+') as f:
        thisd = json.load(f)
        del thisd['impactFactors'][0] # delete the two factors that are still there from copying the files
        del thisd['impactFactors'][0]
        # add new impact factors from master data
        for mli in range(0,DN):
            if dselect[mli] != 1 and TFSel[mli,m] == 1: # add this value as impact/characterisation factor
                CF = CF_generate(mli,dtmr,dnames,duuid,dunit,ei_version_string)
                # add new CF to json file:
                thisd['impactFactors'].append(CF)
        # wrap up and save    
        f.seek(0)        # reset file position to the beginning.
        json.dump(thisd, f, indent=4)
        f.truncate()     # remove remaining part
        f.close()
        
# loop over WF files
for m in range(0,4):
    f_in = os.path.join(tp,'lcia_categories',wf_uuid[m]+'.json')
    with open(f_in, 'r+') as f:
        thisd = json.load(f)
        del thisd['impactFactors'][0] # delete the two factors that are still there from copying the files
        del thisd['impactFactors'][0]
        # add new impact factors from master data
        for mli in range(0,DN):
            if dselect[mli] != 1 and WFSel[mli,m] == 1: # add this value as impact/characterisation factor
                CF = CF_generate(mli,drmi,dnames,duuid,dunit,ei_version_string)
                # add new CF to json file:
                thisd['impactFactors'].append(CF)
        # wrap up and save    
        f.seek(0)        # reset file position to the beginning.
        json.dump(thisd, f, indent=4)
        f.truncate()     # remove remaining part
        f.close()        
            
#%% Sandbox
#

#
#
# The End
#
#
